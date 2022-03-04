from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table
import pathlib

class Conexao_Planilha:


    def __init__(self,disciplina):
        self.disciplina = disciplina

    
    def CriandoArquivo_Excl():
        arquivo_excel = pathlib.Path("output/relatorio.xlsx")
        if arquivo_excel.exists ():
            pass
        else:
            arquivo_excel = Workbook()
            sheet = arquivo_excel.active
            sheet["A1"] = "Codigo"
            sheet["B1"] = "Disciplina"
            sheet["C1"] = "Créditos"
            sheet["D1"] = "C.H"
            sheet["E1"] = "Período"

            arquivo_excel.save("output/relatorio.xlsx")
        
    
''' def submitExcell ():
        code = entry_code()
        discipline = discipline.get()
        cred = cred.get()
        ch = ch.get()
        period = period.get()

        file = load_workbook("relatorio.xlsx")
        sheet = file.active
        sheet.cell(column = 1,row = sheet.max_row+1,value = code)
        sheet.cell(column = 2,row = sheet.max_row,value = discipline)
        sheet.cell(column = 3,row = sheet.max_row,value = cred)
        sheet.cell(column = 4,row = sheet.max_row,value = ch)
        sheet.cell(column = 5,row = sheet.max_row,value = period)
'''

